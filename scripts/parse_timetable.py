from __future__ import annotations

import argparse
import json
import re
from collections import Counter
from datetime import time
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import openpyxl


DAY_MAP = {
    "mon": "Monday",
    "monday": "Monday",
    "tue": "Tuesday",
    "tues": "Tuesday",
    "tuesday": "Tuesday",
    "wed": "Wednesday",
    "wednesday": "Wednesday",
    "thu": "Thursday",
    "thurs": "Thursday",
    "thursday": "Thursday",
    "fri": "Friday",
    "friday": "Friday",
    "sat": "Saturday",
    "saturday": "Saturday",
    "sun": "Sunday",
    "sunday": "Sunday",
}

TEACHER_TITLES = r"(Dr|Mr|Ms|Mrs|Prof|Engr|Sir|Madam|Ns)\.?"
TEACHER_RE = re.compile(rf"\b{TEACHER_TITLES}\b", re.IGNORECASE)


def norm_spaces(s: str) -> str:
    s = (s or "").replace("\n", " ")
    s = re.sub(r"\s+", " ", s).strip()
    return s


def normalize_day(raw: Any) -> Optional[str]:
    if raw is None:
        return None
    s = norm_spaces(str(raw)).lower()
    # sometimes sheet has "Mon " etc
    s = s.replace(".", "")
    if not s:
        return None
    key = s[:3] if len(s) >= 3 else s
    return DAY_MAP.get(s) or DAY_MAP.get(key)


def t_to_str(t: time) -> str:
    return f"{t.hour:02d}:{t.minute:02d}"


def to_minutes(hh: int, mm: int) -> int:
    return hh * 60 + mm


def from_minutes(m: int) -> str:
    hh = (m // 60) % 24
    mm = m % 60
    return f"{hh:02d}:{mm:02d}"


def infer_slot_delta(times: List[time]) -> int:
    if len(times) < 2:
        return 50
    diffs = []
    for a, b in zip(times, times[1:]):
        diffs.append(to_minutes(b.hour, b.minute) - to_minutes(a.hour, a.minute))
    # most common delta, fallback 50
    c = Counter(diffs)
    return c.most_common(1)[0][0] if c else 50


def split_course_teacher(cell_text: str) -> Tuple[Optional[str], Optional[str]]:
    txt = norm_spaces(cell_text)
    if not txt:
        return None, None

    # find first teacher title occurrence and split from there
    m = TEACHER_RE.search(txt)
    if m:
        idx = m.start()
        course = norm_spaces(txt[:idx])
        teacher = norm_spaces(txt[idx:])
        # cleanup title dots
        teacher = teacher.replace("Dr.", "Dr").replace("Mr.", "Mr").replace("Ms.", "Ms").replace("Mrs.", "Mrs")
        return (course or None), (teacher or None)

    # fallback: try split by many spaces (course then teacher)
    parts = re.split(r"\s{2,}", txt)
    if len(parts) >= 2:
        course = norm_spaces(parts[0])
        teacher = norm_spaces(" ".join(parts[1:]))
        return (course or None), (teacher or None)

    return txt, None


def is_section(val: Any) -> bool:
    if val is None:
        return False
    s = norm_spaces(str(val)).upper().replace(" ", "")
    return bool(re.match(r"^BS\d+[A-Z]$", s))


def parse_sheet(ws) -> Dict[str, Any]:
    entries: List[Dict[str, Any]] = []

    r = 1
    max_r = ws.max_row
    max_c = ws.max_column

    while r <= max_r:
        a = ws.cell(r, 1).value
        b = ws.cell(r, 2).value
        c = ws.cell(r, 3).value

        # Detect header row for a day block:
        # Column B = Sections and Column C = No of Students (case-insensitive)
        if (b and str(b).strip().lower() == "sections") and (c and "students" in str(c).strip().lower()):
            day = normalize_day(a)
            if not day:
                r += 1
                continue

            # Extract time columns in this header row
            # NOTE: In the FUUAST sheet, the time headers for CLASSES and ROOMS are often in the SAME ROW
            # and may be continuous columns. The reliable separator is that the time values repeat
            # (e.g., 09:00 appears twice). We split at the first repeated time.
            time_positions = [(col, ws.cell(r, col).value) for col in range(1, max_c + 1)]
            time_positions = [(col, v) for (col, v) in time_positions if isinstance(v, time) and col >= 4]

            if not time_positions:
                r += 1
                continue

            class_seg: Optional[List[Tuple[int, time]]] = None
            room_seg: Optional[List[Tuple[int, time]]] = None

            # 1) Primary split: first repeated time marks start of room segment
            seen_times: set[Tuple[int, int]] = set()
            split_idx: Optional[int] = None
            for i, (col, tv) in enumerate(time_positions):
                key = (tv.hour, tv.minute)
                if key in seen_times:
                    split_idx = i
                    break
                seen_times.add(key)

            if split_idx and split_idx >= 2:
                class_seg = time_positions[:split_idx]
                room_seg = time_positions[split_idx:]
            else:
                # 2) Fallback split by column gaps (older heuristic)
                segs: List[List[Tuple[int, time]]] = []
                current: List[Tuple[int, time]] = []
                last_col = None
                for col, tv in time_positions:
                    if last_col is None or col == last_col + 1:
                        current.append((col, tv))
                    else:
                        if current:
                            segs.append(current)
                        current = [(col, tv)]
                    last_col = col
                if current:
                    segs.append(current)

                if segs:
                    class_seg = segs[0]
                    room_seg = segs[1] if len(segs) > 1 else None

            if not class_seg:
                r += 1
                continue

            times_only = [tv for _, tv in class_seg]
            delta = infer_slot_delta(times_only)

            # Build class start time -> room column mapping.
            # Room headers sometimes differ slightly (e.g., 12:20 vs 12:30), so we match by nearest time (<=20 min).
            room_cols_by_time: Dict[str, int] = {}
            if room_seg:
                room_cands: List[Tuple[int, int]] = []  # (col, minutes)
                for col, tv in room_seg:
                    room_cands.append((col, to_minutes(tv.hour, tv.minute)))

                for col, tv in class_seg:
                    start_str = t_to_str(tv)
                    target_m = to_minutes(tv.hour, tv.minute)
                    best_col = None
                    best_diff = 10**9
                    for rc, rm in room_cands:
                        diff = abs(rm - target_m)
                        if diff < best_diff:
                            best_diff = diff
                            best_col = rc
                    if best_col is not None and best_diff <= 20:
                        room_cols_by_time[start_str] = best_col

            # The class columns are ONLY the class segment
            time_cols = class_seg

            # Now process section rows until next header/day block
            r += 1
            while r <= max_r:
                # Stop if next day header encountered
                b2 = ws.cell(r, 2).value
                c2 = ws.cell(r, 3).value
                if (b2 and str(b2).strip().lower() == "sections") and (c2 and "students" in str(c2).strip().lower()):
                    break

                section_val = ws.cell(r, 2).value
                if is_section(section_val):
                    section = norm_spaces(str(section_val)).upper().replace(" ", "")
                    cap_val = ws.cell(r, 3).value
                    capacity = int(cap_val) if isinstance(cap_val, (int, float)) and cap_val else None

                    for i, (col, tv) in enumerate(time_cols):
                        cell_val = ws.cell(r, col).value
                        if cell_val is None:
                            continue

                        course, teacher = split_course_teacher(str(cell_val))

                        start = t_to_str(tv)
                        if i + 1 < len(time_cols):
                            next_t = time_cols[i + 1][1]
                            end = t_to_str(next_t)
                        else:
                            end_m = to_minutes(tv.hour, tv.minute) + delta
                            end = from_minutes(end_m)

                        room_col = room_cols_by_time.get(start)
                        room = None
                        if room_col:
                            room_val = ws.cell(r, room_col).value
                            if room_val is not None:
                                room = norm_spaces(str(room_val))

                        entries.append({
                            "day": day,
                            "section": section,
                            "capacity": capacity,
                            "start": start,
                            "end": end,
                            "course": course,
                            "teacher": teacher,
                            "room": room,
                        })

                r += 1

            continue

        r += 1

    return {"entries": entries}


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--input", required=True, help="Path to timetable.xlsx")
    ap.add_argument("--output", required=True, help="Path to timetable.json")
    args = ap.parse_args()

    in_path = Path(args.input)
    out_path = Path(args.output)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.load_workbook(in_path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    parsed = parse_sheet(ws)
    payload = {
        "meta": {
            "source_file": in_path.name,
            "sheet": ws.title,
        },
        "entries": parsed["entries"],
    }

    out_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"✅ Wrote {len(payload['entries'])} entries to {out_path}")


if __name__ == "__main__":
    main()
